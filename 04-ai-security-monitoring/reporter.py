import pandas as pd
import json
from detection_engine import AIDetectionEngine

def generate_excel_report(json_log_path="logs/telemetry.json", output_path="logs/SOC_Dashboard.xlsx"):
    detector = AIDetectionEngine()
    logs = []

    # 1. Read raw JSON logs
    with open(json_log_path, "r") as f:
        for line in f:
            logs.append(json.loads(line))

    df = pd.DataFrame(logs)

    # 2. Run Detections and add a "Security_Alerts" column
    df['Alerts'] = df.apply(lambda row: ", ".join(detector.scan_for_threats(row)), axis=1)
    
    # 3. Add Severity Score based on alerts
    def calculate_severity(alert_str):
        if "CRITICAL" in alert_str: return "High"
        if "HIGH" in alert_str: return "Medium"
        if "MEDIUM" in alert_str: return "Low"
        return "Informational"

    df['Severity'] = df['Alerts'].apply(calculate_severity)

    # 4. Export to Excel with formatting
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Security Logs')
        
        # Access the openpyxl workbook to add styling
        workbook = writer.book
        worksheet = writer.sheets['Security Logs']

        # Simple conditional formatting: Highlight Critical/High rows
        from openpyxl.styles import PatternFill
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        for row in range(2, len(df) + 2):  # Skip header
            severity_cell = worksheet.cell(row=row, column=df.columns.get_loc("Severity") + 1)
            if severity_cell.value == "High":
                for col in range(1, len(df.columns) + 1):
                    worksheet.cell(row=row, column=col).fill = red_fill

    print(f"Successfully generated SOC Dashboard at {output_path}")

if __name__ == "__main__":
    generate_excel_report()